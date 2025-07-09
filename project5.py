import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

# === Step 1: Read the input Excel ===
df = pd.read_excel("payroll_guardian_input.xlsx")

# === Step 2: Define compliance rules ===
def validate_pan(pan):
    return isinstance(pan, str) and bool(re.match(r"^[A-Z]{5}[0-9]{4}[A-Z]$", pan))

def validate_email(email):
    return isinstance(email, str) and "@" in email and "." in email

def validate_pf(pf, basic):
    return pf >= 0.12 * basic

def validate_tds(tds):
    return tds >= 5

def validate_doj(doj):
    if isinstance(doj, str):
        doj = datetime.strptime(doj, "%Y-%m-%d")
    return doj <= datetime.today()

def validate_hra(hra):
    return hra < 150000

# === Step 3: Apply compliance logic ===
compliance_col = []
summary = {
    "total": 0, "valid_pan": 0, "valid_email": 0, "valid_pf": 0,
    "valid_tds": 0, "valid_doj": 0, "valid_hra": 0,
    "compliant": 0, "non_compliant": 0
}

for _, row in df.iterrows():
    summary["total"] += 1
    issues = []

    if validate_pan(row["PAN"]): summary["valid_pan"] += 1
    else: issues.append("Invalid PAN")

    if validate_email(row["Email"]): summary["valid_email"] += 1
    else: issues.append("Invalid Email")

    if validate_pf(row["PF Deduction"], row["Basic Salary"]): summary["valid_pf"] += 1
    else: issues.append("PF < 12%")

    if validate_tds(row["TDS (%)"]): summary["valid_tds"] += 1
    else: issues.append("TDS < 5%")

    if validate_doj(row["DOJ"]): summary["valid_doj"] += 1
    else: issues.append("Future DOJ")

    if validate_hra(row["HRA"]): summary["valid_hra"] += 1
    else: issues.append("Unusually high HRA")

    if not issues:
        compliance_col.append("✅ All OK")
        summary["compliant"] += 1
    else:
        compliance_col.append("❌ " + ", ".join(issues))
        summary["non_compliant"] += 1

df["Compliance Status"] = compliance_col

# === Step 4: Save cleaned Excel ===
df.to_excel("payroll_guardian_cleaned.xlsx", index=False)

# === Step 5: Highlight non-compliant rows in Excel ===
wb = load_workbook("payroll_guardian_cleaned.xlsx")
ws = wb.active
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
status_col_index = [cell.value for cell in ws[1]].index("Compliance Status") + 1

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    cell = row[status_col_index - 1]
    if cell.value and "❌" in str(cell.value):
        for c in row:
            c.fill = red_fill

wb.save("payroll_guardian_colored.xlsx")

# === Step 6: Save summary text report ===
summary_text = f"""PAYROLL COMPLIANCE REPORT – Generated on {datetime.today().strftime('%Y-%m-%d %H:%M')}
---------------------------------------------------------
Total Employees Checked: {summary['total']}
✅ Fully Compliant: {summary['compliant']}
❌ Non-Compliant: {summary['non_compliant']}

Breakdown:
- Valid PAN: {summary['valid_pan']}
- Valid Email: {summary['valid_email']}
- PF ≥ 12%: {summary['valid_pf']}
- TDS ≥ 5%: {summary['valid_tds']}
- Valid DOJ: {summary['valid_doj']}
- HRA Sanity OK: {summary['valid_hra']}

Compliance Score: {int((summary['compliant'] / summary['total']) * 100)}%
---------------------------------------------------------
"""

with open("payroll_guardian_summary.txt", "w", encoding="utf-8") as f:
    f.write(summary_text)

# === Step 7: Generate PDF summary report ===
c = canvas.Canvas("payroll_guardian_summary.pdf", pagesize=A4)
width, height = A4

c.setFont("Helvetica-Bold", 14)
c.drawString(50, height - 50, "PAYROLL COMPLIANCE REPORT – July 2025")

c.setFont("Helvetica", 10)
c.drawString(50, height - 80, f"Total Employees Checked: {summary['total']}")
c.drawString(50, height - 100, f"✅ Fully Compliant: {summary['compliant']}")
c.drawString(50, height - 120, f"❌ Non-Compliant: {summary['non_compliant']}")

c.setFont("Helvetica-Bold", 11)
c.drawString(50, height - 150, "Breakdown:")
c.setFont("Helvetica", 10)
c.drawString(70, height - 170, f"- Valid PAN: {summary['valid_pan']}")
c.drawString(70, height - 185, f"- Valid Email: {summary['valid_email']}")
c.drawString(70, height - 200, f"- PF ≥ 12%: {summary['valid_pf']}")
c.drawString(70, height - 215, f"- TDS ≥ 5%: {summary['valid_tds']}")
c.drawString(70, height - 230, f"- Valid DOJ: {summary['valid_doj']}")
c.drawString(70, height - 245, f"- HRA Sanity OK: {summary['valid_hra']}")

score = int((summary["compliant"] / summary["total"]) * 100)
c.setFont("Helvetica-Bold", 12)
c.drawString(50, height - 275, f"Compliance Score: {score}%")

c.setFont("Helvetica-Oblique", 8)
c.drawString(50, height - 300, f"Generated on: {datetime.today().strftime('%Y-%m-%d %H:%M')}")

c.save()

print("✅ All files generated: Cleaned Excel, Highlighted Excel, TXT Summary, PDF Summary.") 

import smtplib
from email.message import EmailMessage

# === Email Configuration ===
sender_email = "soodraghav692@gmail.com"
receiver_email = "anjalichawla19766@gmail.com"
app_password = "phou lsie sktx lvxx"

# Create email
msg = EmailMessage()
msg["Subject"] = "Payroll Compliance Report – July 2025"
msg["From"] = sender_email
msg["To"] = receiver_email
msg.set_content("Hi,\n\nPlease find attached the payroll compliance report and flagged data for July.\n\nRegards,\nPayroll Guardian")

# Attach Excel
with open("payroll_guardian_colored.xlsx", "rb") as f:
    msg.add_attachment(f.read(), maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="payroll_guardian_colored.xlsx")

# Attach PDF
with open("payroll_guardian_summary.pdf", "rb") as f:
    msg.add_attachment(f.read(), maintype="application", subtype="pdf", filename="payroll_guardian_summary.pdf")

# Send via Gmail SMTP
with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
    smtp.login(sender_email, app_password)
    smtp.send_message(msg)

print("✅ Email sent successfully!")


